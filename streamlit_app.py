import streamlit as st
import time
import os
import pandas as pd
from pathlib import Path
import matplotlib.pyplot as plt
import numpy as np
from datetime import date
import openpyxl
from openpyxl import load_workbook
import io
import tempfile # For creating temporary batch file
import atexit # For cleaning up temporary file

# Function definitions moved to the top
def scale_lines(lines, old_area, new_area):
    """
    Scales lines based on the resizing of the working area.
    """
    old_width, old_height = old_area
    new_width, new_height = new_area

    # Scaling factors
    scale_x = new_width / old_width
    scale_y = new_height / old_height

    # Scale the lines
    scaled_lines = [
        ((x1 * scale_x, y1 * scale_y), (x2 * scale_x, y2 * scale_y))
        for ((x1, y1), (x2, y2)) in lines
    ]

    return scaled_lines

def draw_lines(lines, area, title="Lines", colors_list=None):
    """
    Draws lines on a plot.
    """
    width, height = area
    fig, ax = plt.subplots(figsize=(8, 8))
    ax.set_xlim(0, (width * 1.08))
    ax.set_ylim(0, (height * 1.08))
    ax.set_aspect('equal', adjustable='box')
    
    # Draw the working area edges
    ax.plot([0, width, width, 0, 0], [0, 0, height, height, 0], color='black', linestyle='-', linewidth=3)

    default_plot_colors = ['blue', 'green', 'red', 'cyan', 'magenta', 'yellow', 'purple', 'orange', 'brown']
    colors_to_use = colors_list if colors_list is not None and len(colors_list) > 0 else default_plot_colors

    for i, ((x1, y1), (x2, y2)) in enumerate(lines):
        color = colors_to_use[i % len(colors_to_use)]
        ax.plot([x1, x2], [y1, y2], marker="o", color=color)

    # Annotate the width and height of the working area
    ax.annotate(f"Width: {width}", xy=(width / 2, height * 1.02), ha='center', fontsize=10, color='blue')
    ax.annotate(f"Height: {height}", xy=(width * 1.02, height / 2), va='center', rotation=-90, fontsize=10, color='blue')

    ax.set_title(title)
    ax.set_xlabel("Width")
    ax.set_ylabel("Height")
    ax.grid(True)
    
    return fig

# Helper function for styling out-of-bounds coordinates in the DataFrame
def highlight_out_of_bounds_styler(row, new_w, new_h):
    styles = pd.Series('', index=row.index)
    # Check Scaled X start
    if pd.notna(row['Scaled X start']) and (row['Scaled X start'] < 0 or row['Scaled X start'] > new_w):
        styles['Scaled X start'] = 'background-color: orange'
    # Check Scaled Y start
    if pd.notna(row['Scaled Y start']) and (row['Scaled Y start'] < 0 or row['Scaled Y start'] > new_h):
        styles['Scaled Y start'] = 'background-color: orange'
    # Check Scaled X end
    if pd.notna(row['Scaled X end']) and (row['Scaled X end'] < 0 or row['Scaled X end'] > new_w):
        styles['Scaled X end'] = 'background-color: orange'
    # Check Scaled Y end
    if pd.notna(row['Scaled Y end']) and (row['Scaled Y end'] < 0 or row['Scaled Y end'] > new_h):
        styles['Scaled Y end'] = 'background-color: orange'
    return styles

def format_coordinates_to_decimal_places(line, decimals=2):
    """
    Formats the coordinates of a line to the specified number of decimal places.

    Args:
        line (tuple): A tuple of start and end coordinates.
        decimals (int): The number of decimal places.
    
    Returns:
        tuple: The line with coordinates rounded to the specified decimal places.
    """
    start, end = line
    formatted_start = tuple(round(coord, decimals) for coord in start)
    formatted_end = tuple(round(coord, decimals) for coord in end)
    return formatted_start, formatted_end

# Constants and Mappings for Fabricated Sample Exporter
FAB_MATERIALS = ["Silicon", "PET Sheet", "Float Glass"]
FAB_MATERIALS_MAPPING = {"Silicon": "SA00", "PET Sheet": "PB", "Float Glass": "GB"}
FAB_MASTER_IDS = [str(i) for i in range(0, 100)]  # Master Mould choices (0-99 for ID part of name)

# This mapping seems to be for specific, existing master names, which differs from the 0-99 ID.
# For the new tab, we'll use the 0-99 ID for name generation as per generate_sample_name logic.
# The reverse_mapping from the notebook was used to get a descriptive name for the Excel sheet.
# We'll need to decide how to handle this: either use the ID directly in the sheet,
# or if a descriptive name is needed, we might need a different input or mapping.
# For now, master_name_for_excel will be derived from the ID.
FAB_MASTER_NAME_DESCRIPTIVE_MAPPING = {
    0: "Is Master", 1: "PD-SA0002A-JS-A", 2: "PD-SA0002A-KB-A", 3: "PD-SA0002A-KS-A",
    4: "PD-SA0002A-FT-A", 5: "PD-SA0002A-FT-B", 6: "PD-SA0002A-FT-C", 7: "PD-SA0002A-FT-D",
    8: "PD-SA0002A-JS-B", 9: "PD-SA0002B-JS-C", 10: "PD-SA0002A-FT-E", 11: "PD-SA0002B-FT-F",
    12: "PD-SA0002B-FT-G", 13: "PD-SA0000A-JS-A", 14: "PD-SA0002A-FT-H", 15: "PD-SA0002A-FT-I",
    16: "PD-SA0002A-JS-D", 17: "PD-SA0002A-FT-J"
}
# Reverse mapping for descriptive master name (if needed for display/lookup, not directly for name generation from ID)
FAB_REVERSE_MASTER_NAME_MAPPING = {v: k for k, v in FAB_MASTER_NAME_DESCRIPTIVE_MAPPING.items()}


FAB_SALINISATION = [chr(65 + i) for i in range(26)]  # A-Z
FAB_ANTI_STICKING = ["OP-F17G163", "1H,1H,2H,2Hperfluorooctyl-trichlorosilane"]
FAB_ANTI_STICKING_MAPPING = {"OP-F17G163": 1, "1H,1H,2H,2Hperfluorooctyl-trichlorosilane": 2}
FAB_RESIN = ["PS90", "PS380", "OrmoStamp", "UV-PDMS KER-4690 A and B"]
FAB_RESIN_MAPPING = {"PS90": 1, "PS380": 2, "OrmoStamp": 3, "UV-PDMS KER-4690 A and B": 4}
FAB_RESIST = ["OP-PR192", "mr-UVCur26SF", "MM1158A", "SU8", "mr-InkNIL26SF_XP", "OrmoJet_XP"]
FAB_RESIST_MAPPING = {"OP-PR192": 1, "mr-UVCur26SF": 2, "MM1158A": 3, "SU8": 4, "mr-InkNIL26SF_XP": 5, "OrmoJet_XP": 6}
FAB_PRIMER = ["Morphotonics Primer", "OrmoPrime20", "mr-APS1", "OP-APMEX"]
# FAB_PRIMER_MAPPING = {"Morphotonics Primer": 1, "OrmoPrime20": 2, "mr-APS1": 3, "OP-APMEX": 4} # Not used in sample name
FAB_PILLAR_PATTERN = ["Pillar with mesa", "Half Pyramids"]
# FAB_PILLAR_PATTERN_MAPPING = {"Pillar with mesa": 1, "Half Pyramids": 2} # Not used in sample name
FAB_PILLAR_ARRAY = ["5x5", "2x8"]
# FAB_PILLAR_ARRAY_MAPPING = {"5x5": 1, "2x8": 2} # Not used in sample name
FAB_PET = ["VIC Plastics", "HiFi Film", "Sample PET", "PET-PCB without coverslips", "PET-PCB with coverslips", "OPTool PET"]
# FAB_PET_MAPPING = {"VIC Plastics": 1, "HiFi Film": 2, "Sample PET": 3, "PET-PCB without coverslips": 4, "PET-PCB with coverslips": 5, "OPTool PET": 6} # Not used in sample name

FABRICATOR_DEFAULT_VALUES = {
    "Internal Name": "", "Temperature": 22.0, "Pressure": 2.4, "UV": 20.0, "Speed": 200.0,
    "Im_gap": 0.2, "Im_pressure": 2.4, "Del_gap": 0.0, "Del_pressure": 5.5, "Vacuum": 10.0,
    "Pillar Pattern": FAB_PILLAR_PATTERN[0] if FAB_PILLAR_PATTERN else "", 
    "Pillar Array": FAB_PILLAR_ARRAY[0] if FAB_PILLAR_ARRAY else "", 
    "Primer": FAB_PRIMER[0] if FAB_PRIMER else "", 
    "PET": FAB_PET[0] if FAB_PET else "",
    "Metallisation": "False", "Metalised Material": "", "Singulation": "False",
    "Comments": "", "Usability": "False", "UV_Time": 10.0 # Added UV Time, assuming a default
}

# Default lines and parameters for Line Scaling Tab (moved up)
default_lines = [
    ((430, 120), (430, 1000)),  
    ((250, 1000), (250, 120)),  
    ((787, 120), (787, 1000)),  
    ((650, 1000), (650, 120)),  
    ((1100, 120), (1100, 1000)),
    ((1250, 1010), (790, 1010)),
    ((1250, 90), (790, 90)),  
    ((1250, 55), (1250, 1050))
]
default_speed = [58.1, 282.7, 66.9, 138.9, 53.5, 229.7, 229.7, 111.5]
default_t_cycle = [5, 20, 5, 5, 5, 20, 20, 5]
default_t_pulse = [2, 2, 2, 2, 2, 2, 2, 2]
PLOT_COLORS = ['blue', 'green', 'red', 'cyan', 'magenta', 'yellow', 'purple', 'orange', 'brown'] # Also moved related constant

# Function to generate sample name for Fabricated Sample Exporter
def generate_sample_name_fab(materials_fab, master_id_fab, salinisation_fab, anti_sticking_fab, resin_fab, resist_fab, initials_fab, num_samples_fab):
    sample_names = []
    
    materials_mapping_number = FAB_MATERIALS_MAPPING.get(materials_fab)
    if materials_mapping_number is None:
        st.error(f"Selected material '{materials_fab}' is not valid.")
        return []
        
    anti_sticking_mapping_number = FAB_ANTI_STICKING_MAPPING.get(anti_sticking_fab)
    if anti_sticking_mapping_number is None:
        st.error(f"Selected anti-sticking material '{anti_sticking_fab}' is not valid.")
        return []
    
    resin_mapping_number = FAB_RESIN_MAPPING.get(resin_fab)
    if resin_mapping_number is None:
        st.error(f"Selected resin material '{resin_fab}' is not valid.")
        return []

    resist_mapping_number = FAB_RESIST_MAPPING.get(resist_fab)
    if resist_mapping_number is None:
        st.error(f"Selected resist material '{resist_fab}' is not valid.")
        return []
        
    # Resin Name Format from notebook: PD-{materials_mapping_number}{master}0{anti_sticking_mapping_number}{salinisation}-0{resin_mapping_number}-0{resist_mapping_number}-{initials}
    # Ensuring master_id_fab is formatted correctly (e.g., two digits if that's the convention from 'master' list range 0-99)
    # The notebook uses `master` directly which is `str(i) for i in range(0,100)`.
    # If master_id_fab needs to be padded (e.g. 01, 02), this should be handled. Assuming it's used as is for now.
    formatted_master_id_fab = master_id_fab.zfill(2) # Pad with leading zero if single digit
    resin_name_base = f"PD-{materials_mapping_number}{formatted_master_id_fab}0{anti_sticking_mapping_number}{salinisation_fab}-0{resin_mapping_number}-0{resist_mapping_number}-{initials_fab}"
    
    for i in range(num_samples_fab):
        sample_letter = chr(65 + i)
        full_sample_name = f"{resin_name_base}-{sample_letter}"
        sample_names.append(full_sample_name)
    
    return sample_names

# Function to append generated samples to the excel sheet for Fabricated Sample Exporter
def append_sample_data_to_excel_fab(uploaded_file_obj, target_sheet_name, sample_names_fab, internal_name_fab, material_fab, master_name_for_excel_fab, 
                                 ips_name_fab, anti_sticking_fab, resin_fab, anti_sticking2_fab, resist_fab, no_of_prints_fab, 
                                 temperature_fab, pressure_fab, uv_fab, uv_time_fab, speed_fab, im_gap_fab, im_pressure_fab, 
                                 del_gap_fab, del_pressure_fab, vacuum_fab, pillar_pattern_fab, pillar_array_fab, primer_fab, 
                                 pet_fab, metallisation_fab, metalised_material_fab, singulation_fab, comments_fab, usability_fab):
    try:
        # Try to load from staged workbook buffer first
        if st.session_state.fab_staged_workbook_buffer and hasattr(st.session_state.fab_staged_workbook_buffer, 'getvalue') and st.session_state.fab_staged_workbook_buffer.getvalue():
            st.session_state.fab_staged_workbook_buffer.seek(0)
            wb = openpyxl.load_workbook(st.session_state.fab_staged_workbook_buffer)
        elif uploaded_file_obj:
            wb = openpyxl.load_workbook(uploaded_file_obj)
        else:
            wb = openpyxl.Workbook()

        if target_sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=target_sheet_name)
        else:
            ws = wb[target_sheet_name]
        
        # If creating a brand new workbook and the default "Sheet" exists and is empty, and target is not "Sheet"
        # or if "Sheet" exists, is empty, and target is "Sheet", use it. Otherwise, if it exists and has data, remove it
        # if we intend to create a new sheet with a different name.
        # This logic is complex, simplifying for now: if creating new and target is not "Sheet", remove "Sheet" if it exists empty.
        if not uploaded_file_obj and not (st.session_state.fab_staged_workbook_buffer and st.session_state.fab_staged_workbook_buffer.getvalue()) and "Sheet" in wb.sheetnames and target_sheet_name != "Sheet":
            # Check if the default "Sheet" is empty before removing
            default_sheet_check = wb["Sheet"]
            if not any(default_sheet_check.iter_rows(values_only=True)):
                 wb.remove(default_sheet_check)
            # Re-create or get the target sheet if it was the default one and got removed above.
            if target_sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=target_sheet_name)
            else:
                ws = wb[target_sheet_name]


        # Find the true last row with data
        true_last_data_row = 0
        if ws.max_row > 0: # Only check if there are any rows reported by max_row
            for row_num in range(ws.max_row, 0, -1):
                row_has_data = False
                for col_num in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value is not None and str(cell_value).strip(): # Check for non-empty, non-whitespace values
                        row_has_data = True
                        break
                if row_has_data:
                    true_last_data_row = row_num
                    break
        
        next_row_to_write = true_last_data_row + 1
        
        # Write headers if the sheet is effectively empty (true_last_data_row is 0)
        if true_last_data_row == 0:
            headers = [
                "Sample Name", "Internal Name", "Material", "Master Name", "Date", 
                "IPS Name", "Anti Sticking", "Resin", "Anti Sticking 2", "Resist", 
                "No of Prints", "Temperature", "Pressure", "UV", "UV Time", "Speed", 
                "Im_gap", "Im_pressure", "Del_gap", "Del_pressure", "Vacuum", 
                "Pillar Pattern", "Pillar Array", "Primer", "PET", "Metallisation",
                "Metalised Material", "Singulation", "Comments", "Usability"
            ] 
            for col_idx, header_title in enumerate(headers):
                ws.cell(row=next_row_to_write, column=col_idx + 1).value = header_title
            next_row_to_write += 1 # Move to next row for actual data
       
        # Write the actual sample data rows
        for sample_name_fab_val in sample_names_fab:
            today = date.today()
            formatted_date = today.strftime("%d/%m/%Y")
            sample_data_list = [
                sample_name_fab_val, internal_name_fab, material_fab, master_name_for_excel_fab, formatted_date, 
                ips_name_fab, anti_sticking_fab, resin_fab, anti_sticking2_fab, resist_fab, no_of_prints_fab, 
                temperature_fab, pressure_fab, uv_fab, uv_time_fab, speed_fab, im_gap_fab, im_pressure_fab, 
                del_gap_fab, del_pressure_fab, vacuum_fab, pillar_pattern_fab, pillar_array_fab, primer_fab,
                pet_fab, metallisation_fab, metalised_material_fab, singulation_fab, comments_fab, usability_fab
            ]
            for col_idx, cell_value in enumerate(sample_data_list):
                ws.cell(row=next_row_to_write, column=col_idx + 1).value = cell_value
            next_row_to_write += 1 # Increment for the next sample
        
        return wb

    except Exception as e:
        st.error(f"Error processing Excel data: {str(e)}")
        return None

# Helper function to load sheet names and preview for Fabricated Sample Exporter
def update_fab_sheet_data(clear_all=False):
    if clear_all:
        st.session_state.fab_excel_sheets_options = []
        st.session_state.fab_selected_sheet_name = None
        st.session_state.fab_df_preview = None
        # Do not clear fab_uploaded_excel_file here as it's managed by the uploader widget directly
        # and its on_change callback handles clearing sheets if the file is removed.
        return

    uploaded_file = st.session_state.fab_uploaded_excel_file # This is an UploadedFile object or None

    if uploaded_file:
        try:
            excel_file_data = pd.ExcelFile(uploaded_file) # pd.ExcelFile can handle UploadedFile objects
            sheet_names = excel_file_data.sheet_names
            st.session_state.fab_excel_sheets_options = sheet_names
            
            if sheet_names:
                if st.session_state.fab_selected_sheet_name not in sheet_names:
                    st.session_state.fab_selected_sheet_name = sheet_names[0]
                
                if st.session_state.fab_selected_sheet_name:
                    st.session_state.fab_df_preview = pd.read_excel(excel_file_data, sheet_name=st.session_state.fab_selected_sheet_name)
                else:
                    st.session_state.fab_df_preview = None
            else:
                st.session_state.fab_excel_sheets_options = []
                st.session_state.fab_selected_sheet_name = None
                st.session_state.fab_df_preview = None
                st.warning(f"No sheets found in uploaded file '{uploaded_file.name}'.")

        except Exception as e:
            update_fab_sheet_data(clear_all=True) # Clear everything on error
            st.error(f"An unexpected error occurred while reading uploaded Excel file '{uploaded_file.name}': {e}")
    else: # No file uploaded
        update_fab_sheet_data(clear_all=True) # Clear sheet options and preview if no file is present
        # No error/warning here as it's normal not to have a file initially.

# Configure the page
st.set_page_config(
    page_title="Advanced Manufacturing Control Panel",
    page_icon="🏭",
    layout="wide"
)

# Initialize session state for file management
if 'selected_files' not in st.session_state:
    st.session_state.selected_files = []
if 'current_directory' not in st.session_state:
    st.session_state.current_directory = os.path.expanduser("~")
# Session state for Data Structure Creator (Tab 4) output location browser
if 'save_location_val' not in st.session_state:
    st.session_state.save_location_val = os.path.expanduser("~")
if 'output_location_browser_path' not in st.session_state:
    st.session_state.output_location_browser_path = st.session_state.save_location_val 
if 'show_dir_browser_tab4' not in st.session_state: 
    st.session_state.show_dir_browser_tab4 = False
# Session state for Line Scaling (Tab 5) - Now uses pre-defined defaults
if 'lines' not in st.session_state:
    st.session_state.lines = default_lines
    st.session_state.speed = default_speed
    st.session_state.t_cycle = default_t_cycle
    st.session_state.t_pulse = default_t_pulse

# Session state for Fabricated Sample Exporter (New Tab)
if 'fab_uploaded_excel_file' not in st.session_state: # For st.file_uploader object
    st.session_state.fab_uploaded_excel_file = None
if 'fab_target_save_path' not in st.session_state: # For the text_input specifying where to save
    st.session_state.fab_target_save_path = ""
if 'fab_selected_sheet_name' not in st.session_state:
    st.session_state.fab_selected_sheet_name = None
if 'fab_excel_sheets_options' not in st.session_state:
    st.session_state.fab_excel_sheets_options = []
if 'fab_df_preview' not in st.session_state:
    st.session_state.fab_df_preview = None

# Session state for managing batches of fabricated samples
if 'fab_staged_workbook_buffer' not in st.session_state: # Stores BytesIO of the workbook being built
    st.session_state.fab_staged_workbook_buffer = None
if 'fab_staged_sample_names' not in st.session_state: # Stores list of names in the current batch
    st.session_state.fab_staged_sample_names = []

# Mock data for machine status - Initialize in session state
if 'machine_status_data' not in st.session_state:
    st.session_state.machine_status_data = {
        'machines': [
            {'id': 1, 'name': 'Primer', 'status': 'Idle', 'uptime': '0h', 'comments': ''},
            {'id': 2, 'name': 'Coater', 'status': 'Idle', 'uptime': '0h', 'comments': ''},
            {'id': 3, 'name': 'Nanoimprint Lithography', 'status': 'Idle', 'uptime': '0h', 'comments': ''},
            {'id': 4, 'name': 'DRIE (ANFF)', 'status': 'Maintenance', 'uptime': '0h', 'comments': 'Needs calibration'},
            {'id': 5, 'name': 'Dicer (ANFF)', 'status': 'Idle', 'uptime': '0h', 'comments': ''},
            {'id': 6, 'name': 'SEM (ANFF)', 'status': 'Running', 'uptime': '1h 15m', 'comments': 'Imaging new samples'}
        ]
    }

# Session state for Rclone Downloader Tab
if 'rclone_remote_name' not in st.session_state:
    st.session_state.rclone_remote_name = ""
if 'rclone_source_path' not in st.session_state:
    st.session_state.rclone_source_path = ""
if 'rclone_local_destination' not in st.session_state:
    st.session_state.rclone_local_destination = os.path.expanduser("~") # Default to home directory
if 'rclone_command_output' not in st.session_state:
    st.session_state.rclone_command_output = ""
if 'rclone_last_run_command' not in st.session_state:
    st.session_state.rclone_last_run_command = ""
if 'rclone_is_running' not in st.session_state: # To control spinner
    st.session_state.rclone_is_running = False
if 'rclone_temp_batch_file_path' not in st.session_state: # To store path of temp .bat file
    st.session_state.rclone_temp_batch_file_path = None

# Session state for Rclone Local Destination Browser
if 'rclone_show_local_dest_browser' not in st.session_state:
    st.session_state.rclone_show_local_dest_browser = False
if 'rclone_local_dest_browser_path' not in st.session_state:
    st.session_state.rclone_local_dest_browser_path = st.session_state.rclone_local_destination # Initialize with current dest

if 'rclone_exe_path' not in st.session_state: # For user to specify rclone.exe location
    st.session_state.rclone_exe_path = "rclone" # Default to 'rclone', assuming it's in PATH

# Remove old python subprocess state (if they exist from previous versions)
if 'rclone_python_execute_request' in st.session_state:
    del st.session_state['rclone_python_execute_request']
if 'rclone_python_code_to_run' in st.session_state:
    del st.session_state['rclone_python_code_to_run']

# Function to clean up the temporary batch file
def cleanup_rclone_batch_file():
    if st.session_state.get('rclone_temp_batch_file_path') and os.path.exists(st.session_state.rclone_temp_batch_file_path):
        try:
            os.remove(st.session_state.rclone_temp_batch_file_path)
            # st.write(f"Cleaned up {st.session_state.rclone_temp_batch_file_path}") # For debugging
            st.session_state.rclone_temp_batch_file_path = None
        except Exception as e:
            # st.write(f"Error cleaning up batch file: {e}") # For debugging
            pass # Fail silently on cleanup

# Register cleanup function to be called on script exit
atexit.register(cleanup_rclone_batch_file)

# Check for rclone command result (now from run_terminal_cmd) and update UI state
if 'tool_run_terminal_cmd_result' in st.session_state and st.session_state.tool_run_terminal_cmd_result is not None:
    terminal_output = st.session_state.tool_run_terminal_cmd_result
    
    stdout = terminal_output.get('stdout', '')
    stderr = terminal_output.get('stderr', '')
    # exit_code = terminal_output.get('exit_code', 'N/A') # Assuming exit_code is part of the result

    output_message = f"Rclone command execution via batch file finished.\n"
    # output_message += f"Exit Code: {exit_code}\n" # If available
    output_message += f"--- stdout ---\n{stdout}\n"
    output_message += f"--- stderr ---\n{stderr}"
    
    st.session_state.rclone_command_output = output_message
    st.session_state.rclone_is_running = False # Stop spinner
    
    cleanup_rclone_batch_file() # Clean up the temp file now that execution is done

    del st.session_state.tool_run_terminal_cmd_result # Clear the result once processed
    st.rerun()

# Function to rename files
def rename_files_in_folder(folder_path, old_string, new_string, prefix_string):
    if not os.path.exists(folder_path):
        st.error(f"The folder {folder_path} does not exist.")
        return []
    
    renamed_files = []
    for filename in os.listdir(folder_path):
        old_file_path = os.path.join(folder_path, filename)
        if os.path.isfile(old_file_path):
            new_filename = filename
            
            # Apply old_string to new_string replacement if old_string is provided
            if old_string:
                new_filename = new_filename.replace(old_string, new_string)

            # Prepend the prefix_string if provided
            if prefix_string:
                new_filename = f"{prefix_string}{new_filename}".replace(' ', '_').replace('_-_', '_')
            else:
                new_filename = new_filename.replace(' ', '_').replace('_-_', '_')

            new_file_path = os.path.join(folder_path, new_filename)

            if old_file_path != new_file_path:
                try:
                    os.rename(old_file_path, new_file_path)
                    renamed_files.append((filename, new_filename))
                except Exception as e:
                    st.error(f"Error renaming {filename}: {str(e)}")
    
    return renamed_files

# Function to save Excel rows as CSV files
def save_rows_as_csv(df, output_folder, start_row, end_row):
    try:
        # Adjust for zero-indexing
        start_row -= 1
        end_row -= 1

        # Ensure the row range is valid
        if start_row < 0 or end_row >= len(df) or start_row > end_row:
            st.error("Please enter a valid row range.")
            return []

        saved_files = []
        # Iterate over the specified rows in the DataFrame
        for index in range(start_row, end_row + 1):
            row = df.iloc[index]

            # Check if the row is completely blank
            if row.isnull().all():
                continue  # Skip this row if it's blank

            # Get the name from the first column
            file_name = str(row[0])  # Assuming the name is in the first column
            
            # Create a new DataFrame with the header and the current row
            row_df = pd.DataFrame([row.values], columns=df.columns)

            # Save the row as a CSV file, ensuring the file name is valid
            if pd.notna(file_name) and file_name:  # Ensure the file name is not NaN or empty
                file_path = os.path.join(output_folder, f"{file_name}.csv")
                row_df.to_csv(file_path, index=False)
                saved_files.append(file_name)
            else:
                st.warning(f"Row {index + 1} has an invalid name; skipping.")

        return saved_files

    except Exception as e:
        st.error(f"Error: {str(e)}")
        return []

def list_directory_contents(path):
    try:
        # Get directories and files
        items = list(Path(path).glob('*'))
        directories = [item for item in items if item.is_dir()]
        files = [item for item in items if item.is_file()]
        
        # Sort alphabetically
        directories.sort(key=lambda x: x.name.lower())
        files.sort(key=lambda x: x.name.lower())
        
        return directories, files
    except Exception as e:
        st.error(f"Error accessing directory: {str(e)}")
        return [], []

def get_file_info(file_path):
    try:
        stats = os.stat(file_path)
        size = stats.st_size
        modified = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(stats.st_mtime))
        
        # Convert size to readable format
        if size < 1024:
            size_str = f"{size} B"
        elif size < 1024 * 1024:
            size_str = f"{size/1024:.1f} KB"
        else:
            size_str = f"{size/(1024*1024):.1f} MB"
            
        return size_str, modified
    except:
        return "N/A", "N/A"

# Function to create folders based on Sample ID and save CSV
def create_folders_for_csv(csv_file, file_name, row_df, save_location, fabrication_checked, inspection_checked):
    created_folders = []
    try:
        # Folder paths based on the Sample ID
        run_folder = os.path.join(save_location, f"Run={file_name}")
        stage_folder = os.path.join(run_folder, "Stage=source_data")
        
        # Create the "Stage=source_data" folder
        os.makedirs(stage_folder, exist_ok=True)
        created_folders.append(stage_folder)

        # Create folders based on options
        if fabrication_checked:
            fabrication_folder = os.path.join(run_folder, "Stage=source_data", "Modality=record_manufacture")
            os.makedirs(fabrication_folder, exist_ok=True)
            created_folders.append(fabrication_folder)
            # Save CSV in the "record_manufacture" folder
            csv_file_path = os.path.join(fabrication_folder, csv_file)
            row_df.to_csv(csv_file_path, index=False)

        if inspection_checked:
            # Create all modality folders except "Modality=record_manufacture"
            modalities = [
                "Modality=optical_image",
                "Modality=sem_c_0deg",
                "Modality=sem_c_high_angle",
                "Modality=sem_c_medium_angle",
                "Modality=sem_p_0deg",
                "Modality=sem_p_high_angle",
                "Modality=sem_p_medium_angle"
            ]
            
            for modality in modalities:
                modality_folder = os.path.join(stage_folder, modality)
                os.makedirs(modality_folder, exist_ok=True)
                created_folders.append(modality_folder)
            
            # Save the CSV in the stage folder if inspection_checked is true
            # This is the key part for inspection CSV
            csv_file_path_inspection = os.path.join(stage_folder, csv_file)
            row_df.to_csv(csv_file_path_inspection, index=False)
        
        return True, created_folders
    except Exception as e:
        return False, str(e)

# Title and description
st.title("Advanced Manufacturing Control Panel")
st.markdown("Monitor and control manufacturing equipment from anywhere")

# Create tabs for different functionalities
tab1, tab2, tab_fab_exporter, tab_rclone, tab3, tab4, tab5 = st.tabs([
    "Equipment Dashboard", 
    "File Management",
    "Fabricated Sample Exporter",
    "SharePoint Download (Rclone)",
    "Excel Row Exporter", 
    "Data Structure Creator",
    "Line Scaling"
])

# Tab 1: Equipment Dashboard
with tab1:
    # Create columns for the dashboard
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Equipment Status")
        # Display each machine's status in a card
        for machine in st.session_state.machine_status_data['machines']:
            with st.container():
                st.markdown(f"""
                <div style='padding: 10px; border: 1px solid #ddd; border-radius: 5px; margin: 10px 0;'>
                    <h3>{machine['name']}</h3>
                    <p>Status: <span style='color: {'green' if machine['status'] == 'Running' else 'orange' if machine['status'] == 'Idle' else 'red'};'>
                        {machine['status']}</span></p>
                    <p>Uptime: {machine['uptime']}</p>
                    <p>Comments: {machine['comments'] if machine['comments'] else 'N/A'}</p>
                </div>
                """, unsafe_allow_html=True)

    with col2:
        st.subheader("Machine Controls")
        # Control panel for each machine
        for idx, machine in enumerate(st.session_state.machine_status_data['machines']):
            with st.expander(f"Control {machine['name']}"):
                # Status control buttons
                col_start, col_stop, col_maint = st.columns(3)
                with col_start:
                    if st.button(f"Set Running", key=f"start_{machine['id']}"):
                        st.session_state.machine_status_data['machines'][idx]['status'] = 'Running'
                        # Potentially update uptime here if logic is added
                        st.success(f"{machine['name']} status set to Running")
                        st.rerun()
                with col_stop:
                    if st.button(f"Set Idle", key=f"stop_{machine['id']}"):
                        st.session_state.machine_status_data['machines'][idx]['status'] = 'Idle'
                        st.warning(f"{machine['name']} status set to Idle")
                        st.rerun()
                with col_maint:
                    if st.button(f"Set Maintenance", key=f"maint_{machine['id']}"):
                        st.session_state.machine_status_data['machines'][idx]['status'] = 'Maintenance'
                        st.info(f"{machine['name']} status set to Maintenance")
                        st.rerun()
                
                # Comment input
                # Use a unique key for the text_area and update session_state based on this key's value
                comment_key = f"comment_input_{machine['id']}"
                # Initialize the text_area with the current comment
                current_comment = st.session_state.machine_status_data['machines'][idx]['comments']
                new_comment = st.text_area("Add/Edit Comment:", value=current_comment, key=comment_key, height=100)
                
                # Update the comment in session_state if it has changed
                if new_comment != current_comment:
                    st.session_state.machine_status_data['machines'][idx]['comments'] = new_comment
                    st.rerun() # Rerun to reflect comment update immediately in status display

    # System metrics
    st.subheader("System Metrics")
    metrics_col1, metrics_col2, metrics_col3 = st.columns(3)

    with metrics_col1:
        st.metric("Total Uptime", "21h 15m")
    with metrics_col2:
        st.metric("Active Machines", "1/3")
    with metrics_col3:
        st.metric("Efficiency", "85%", "+2%")

# Tab 2: File Management
with tab2:
    st.subheader("File Renaming Tool")

    # Helper function to update current directory and select all files
    def fm_update_current_directory_and_select_all(new_path_str):
        # Resolve to an absolute path and normalize
        try:
            resolved_path = str(Path(new_path_str).resolve())
        except Exception as e:
            st.error(f"Error resolving path '{new_path_str}': {e}")
            return False

        if os.path.isdir(resolved_path):
            st.session_state.current_directory = resolved_path
            st.session_state.selected_files.clear()
            _, files_in_new_dir = list_directory_contents(st.session_state.current_directory)
            st.session_state.selected_files.extend([str(f.resolve()) for f in files_in_new_dir])
            # Update the text input field to reflect the new path
            st.session_state.fm_path_input_val = resolved_path 
            return True
        else:
            st.error(f"Invalid or inaccessible directory: {resolved_path}")
            return False

    # Breadcrumbs Navigation
    st.markdown("**Navigate Path:**")
    current_path_obj = Path(st.session_state.current_directory)
    parts = list(current_path_obj.parts)
    breadcrumbs_cols = st.columns(len(parts) + (len(parts) -1) ) # Columns for parts and separators
    
    path_so_far = ""
    for i, part in enumerate(parts):
        if i == 0: # For drive letter like C:\
            path_so_far = part
        else:
            path_so_far = os.path.join(path_so_far, part)
        
        with breadcrumbs_cols[i*2]:
            if i < len(parts) - 1: # Not the last part
                if st.button(part, key=f"breadcrumb_{path_so_far}", help=f"Go to {path_so_far}"):
                    if fm_update_current_directory_and_select_all(path_so_far):
                        st.rerun()
            else: # Last part (current directory name)
                st.markdown(f"**{part}**") # Display last part as bold text, not button
        
        if i < len(parts) - 1:
            with breadcrumbs_cols[i*2 + 1]:
                st.markdown(">")
    st.markdown("---") # Visual separator

    # Directory navigation input
    # Use a temporary variable for the text input to avoid immediate state change on typing
    if 'fm_path_input_val' not in st.session_state:
        st.session_state.fm_path_input_val = st.session_state.current_directory
    
    # Update fm_path_input_val if current_directory changes elsewhere (e.g. up button)
    if st.session_state.fm_path_input_val != st.session_state.current_directory:
        st.session_state.fm_path_input_val = st.session_state.current_directory

    fm_path_input_text = st.text_input(
        "Current Directory Path:", 
        value=st.session_state.fm_path_input_val, 
        key="fm_path_input_field",
        on_change=lambda: setattr(st.session_state, 'fm_path_input_val', st.session_state.fm_path_input_field) # Update temp var on change
    )

    col_nav_buttons1, col_nav_buttons2 = st.columns(2)
    with col_nav_buttons1:
        if st.button("Go to Path", key="fm_go_to_path_btn"):
            # Use the value from the temporary session state variable that reflects the text_input's current content
            if fm_update_current_directory_and_select_all(st.session_state.fm_path_input_val):
                st.rerun()
    
    with col_nav_buttons2:
        if st.button("⬆️ Up One Level", key="fm_up_one_level_btn"):
            parent = str(Path(st.session_state.current_directory).parent)
            if parent != st.session_state.current_directory: # Avoid getting stuck if already at root or similar
                if fm_update_current_directory_and_select_all(parent):
                    st.session_state.fm_path_input_val = parent # Update text input to reflect new path
                    st.rerun()
            else:
                st.info("Already at the top level or cannot go further up.")
    
    # List directories and files
    directories, files = list_directory_contents(st.session_state.current_directory)
    
    # Show directories
    if directories:
        st.subheader("📁 Directories")
        for dir_item in directories: 
            dir_path_str = str(dir_item.resolve()) 
            display_key = f"dir_{str(dir_item)}_{dir_item.name}"
            if st.button(f"📁 {dir_item.name}", key=display_key):
                if fm_update_current_directory_and_select_all(dir_path_str):
                    st.session_state.fm_path_input_val = dir_path_str 
                    st.rerun()
    
    # File filter - Moved before the conditional display of files
    file_filter = st.text_input("🔍 Filter files (leave empty to show all):", "", key="fm_file_filter")

    # Show files with selection
    if files:
        st.subheader("📄 Files")
        
        # Filter files based on the input
        filtered_files = [
            file for file in files 
            if not file_filter or file_filter.lower() in file.name.lower()
        ]

        if filtered_files: # Only show buttons and table if there are files after filtering
            # Select All / Deselect All buttons for VISIBLE files
            col_sel_buttons1, col_sel_buttons2 = st.columns(2)
            with col_sel_buttons1:
                if st.button("Select All Visible Files", key="fm_select_all_visible_btn"):
                    for file_obj in filtered_files:
                        resolved_f_path = str(file_obj.resolve())
                        if resolved_f_path not in st.session_state.selected_files:
                            st.session_state.selected_files.append(resolved_f_path)
                    st.rerun()
            with col_sel_buttons2:
                if st.button("Deselect All Files", key="fm_deselect_all_btn"):
                    st.session_state.selected_files.clear()
                    st.rerun()

            # Create a table-like display for files
            header_cols = st.columns([0.5, 2, 1, 1])
            with header_cols[0]:
                st.write("Select")
            with header_cols[1]:
                st.write("File Name")
            with header_cols[2]:
                st.write("Size")
            with header_cols[3]:
                st.write("Modified")
            
            for file_obj in filtered_files: 
                size, modified = get_file_info(file_obj) 
                resolved_file_path_str = str(file_obj.resolve())
                
                row_cols = st.columns([0.5, 2, 1, 1])
                with row_cols[0]:
                    is_selected = st.checkbox(label=f"Select file {file_obj.name}", key=f"select_{resolved_file_path_str}", value=resolved_file_path_str in st.session_state.selected_files, label_visibility="collapsed")
                    if is_selected:
                        if resolved_file_path_str not in st.session_state.selected_files:
                            st.session_state.selected_files.append(resolved_file_path_str)
                    else:
                        if resolved_file_path_str in st.session_state.selected_files:
                            st.session_state.selected_files.remove(resolved_file_path_str)
                with row_cols[1]:
                    st.write(file_obj.name)
                with row_cols[2]:
                    st.write(size)
                with row_cols[3]:
                    st.write(modified)
        else: # Files exist in the directory, but none match the current filter
            st.info("No files match the current filter.")
            
    else: # No files in the directory at all
        st.info("No files found in this directory.")

    # File renaming options (ensure this uses resolved paths if it interacts with selected_files)
    if st.session_state.selected_files:
        st.subheader("Rename Selected Files")
        st.write(f"Selected {len(st.session_state.selected_files)} files")
        
        col1, col2 = st.columns(2)
        with col1:
            old_string = st.text_input("Text to Replace", placeholder="Leave blank to skip")
            new_string = st.text_input("Replace With", placeholder="New text")
        with col2:
            prefix_string = st.text_input("Add Prefix", placeholder="Leave blank to skip")
            
        if st.button("Preview Changes"):
            st.write("Preview of changes:")
            for file_path in st.session_state.selected_files:
                file = Path(file_path)
                new_filename = file.name
                if old_string:
                    new_filename = new_filename.replace(old_string, new_string)
                if prefix_string:
                    new_filename = f"{prefix_string}{new_filename}"
                new_filename = new_filename.replace(' ', '_').replace('_-_', '_')
                
                if new_filename != file.name:
                    st.text(f"'{file.name}' → '{new_filename}'")
                
        if st.button("Rename Selected Files"):
            renamed_count = 0
            for file_path in st.session_state.selected_files:
                file = Path(file_path)
                new_filename = file.name
                if old_string:
                    new_filename = new_filename.replace(old_string, new_string)
                if prefix_string:
                    new_filename = f"{prefix_string}{new_filename}"
                new_filename = new_filename.replace(' ', '_').replace('_-_', '_')
                
                if new_filename != file.name:
                    try:
                        new_path = file.parent / new_filename
                        os.rename(file, new_path)
                        renamed_count += 1
                    except Exception as e:
                        st.error(f"Error renaming {file.name}: {str(e)}")
            
            if renamed_count > 0:
                st.success(f"Successfully renamed {renamed_count} files!")
                st.session_state.selected_files = []
                st.rerun()
            else:
                st.info("No files were renamed.")

# New Tab: Fabricated Sample Exporter
with tab_fab_exporter:
    st.subheader("Fabricated Sample Exporter")
    st.caption("Enter sample information to generate names and append to an Excel spreadsheet.")

    # --- New File Uploader and Target Save Path ---
    def on_fab_file_upload_change():
        # This callback is triggered when a new file is uploaded or removed.
        # st.session_state.fab_uploaded_excel_file_widget will hold the UploadedFile object or None.
        st.session_state.fab_uploaded_excel_file = st.session_state.fab_uploaded_excel_file_widget
        if st.session_state.fab_uploaded_excel_file:
            # If a file is uploaded, suggest its name for the save path, but only if user hasn't typed something already
            # or if the current save path is empty or matches a previous suggestion from an old file.
            # This check avoids overwriting a user-deliberately-typed path just because they re-uploaded the same file.
            # A more sophisticated check might be needed if we want to track if fab_target_save_path was auto-filled.
            if not st.session_state.fab_target_save_path or \
               (hasattr(st.session_state, '_previous_uploaded_filename') and st.session_state.fab_target_save_path == st.session_state._previous_uploaded_filename) :
                st.session_state.fab_target_save_path = st.session_state.fab_uploaded_excel_file.name
            st.session_state._previous_uploaded_filename = st.session_state.fab_uploaded_excel_file.name # Store for next check
        else:
            # If file is removed, we could clear the target save path or leave it.
            # Clearing might be annoying if user just accidentally removed and re-uploads.
            # For now, let's leave it but clear sheet data.
            st.session_state._previous_uploaded_filename = "" # Clear previous name
        update_fab_sheet_data() # Update sheets based on new upload state

    st.file_uploader(
        "Upload Excel File (optional, to select sheet from or use as name template)",
        type=['xlsx', 'xls'],
        key='fab_uploaded_excel_file_widget', # This key gets the UploadedFile object
        on_change=on_fab_file_upload_change,
        accept_multiple_files=False
    )

    st.session_state.fab_target_save_path = st.text_input(
        "Target Save File Path (.xlsx)", 
        value=st.session_state.fab_target_save_path, 
        key="fab_target_save_path_text_input",
        placeholder="e.g., ./output/my_generated_samples.xlsx"
    )
    st.session_state.fab_target_save_path = st.session_state.fab_target_save_path_text_input # ensure session state is updated from text input if changed
    # --- End New File Uploader ---


    # --- Target Sheet Name and Preview ---
    # if not st.session_state.fab_excel_file_path_val or not (st.session_state.fab_excel_file_path_val.endswith('.xlsx') or st.session_state.fab_excel_file_path_val.endswith('.xls')):
    # Modified condition to check the new fab_uploaded_excel_file state
    if not st.session_state.fab_uploaded_excel_file:
        st.info("Upload an Excel file above to see sheet options and preview, or ensure the Target Save File Path is set if creating a new file.")
        # Ensure sheet related states are cleared if no file is uploaded
        if st.session_state.fab_excel_sheets_options or st.session_state.fab_df_preview is not None:
            update_fab_sheet_data(clear_all=True)

    else: # An Excel file is uploaded
        # update_fab_sheet_data() should have been called by on_fab_file_upload_change
        # So, fab_excel_sheets_options and fab_df_preview should be populated if file is valid.
        
        if st.session_state.fab_excel_sheets_options:
            def fab_update_preview_on_sheet_select():
                # The selectbox's value is already in st.session_state.fab_selected_sheet_name due to `key=`
                update_fab_sheet_data() # This will re-evaluate and load the correct preview based on the new sheet selection from the uploaded file

            st.selectbox(
                "Target Sheet Name (from uploaded file):",
                options=st.session_state.fab_excel_sheets_options,
                key='fab_selected_sheet_name', 
                on_change=fab_update_preview_on_sheet_select,
                index=st.session_state.fab_excel_sheets_options.index(st.session_state.fab_selected_sheet_name) if st.session_state.fab_selected_sheet_name in st.session_state.fab_excel_sheets_options else 0
            )
            if st.session_state.fab_df_preview is not None:
                st.write("Preview of selected sheet (first 5 rows from uploaded file):")
                st.dataframe(st.session_state.fab_df_preview.head())
            elif st.session_state.fab_selected_sheet_name: 
                st.warning(f"Could not load preview for sheet '{st.session_state.fab_selected_sheet_name}' from the uploaded file.")
        
        # If no sheets were found in the uploaded file, update_fab_sheet_data would have shown a warning.
        # We don't need an explicit else here for that case again, as the selectbox just won't appear.

    # If no file is uploaded, user might still want to type a new sheet name for a new file.
    # We need a way for the user to specify the sheet name if no file is uploaded OR if they want to override.
    # Let's make fab_selected_sheet_name editable also via a text input if no file is uploaded or if user wants to override.

    # Allow specifying sheet name manually, especially if no file is uploaded or to create a new sheet.
    current_sheet_name_for_input = st.session_state.fab_selected_sheet_name if st.session_state.fab_selected_sheet_name else "Sheet1"
    
    def on_manual_sheet_name_change():
        st.session_state.fab_selected_sheet_name = st.session_state.fab_manual_sheet_name_input
        # No preview update here as this is for manual entry, potentially for a new file/sheet.

    st.text_input(
        "Target Sheet Name (if creating new, or to override selection from uploaded file):", 
        value=current_sheet_name_for_input, 
        key="fab_manual_sheet_name_input",
        on_change=on_manual_sheet_name_change
    )

    st.markdown("---")
    st.subheader("Sample Base Information")
    
    col_fab1, col_fab2 = st.columns(2)
    with col_fab1:
        fab_material = st.selectbox("Material:", FAB_MATERIALS, key="fab_material")
        fab_master_id = st.selectbox("Master ID (for name generation):", FAB_MASTER_IDS, index=0, key="fab_master_id") # e.g. "0", "1"... "99"
        fab_salinisation = st.selectbox("Salinisation (Letter Code):", FAB_SALINISATION, key="fab_salinisation")
        fab_anti_sticking = st.selectbox("Anti-sticking Agent:", FAB_ANTI_STICKING, key="fab_anti_sticking")
    with col_fab2:
        fab_resin = st.selectbox("Resin:", FAB_RESIN, key="fab_resin")
        fab_resist = st.selectbox("Resist:", FAB_RESIST, key="fab_resist")
        fab_initials = st.text_input("Initials (e.g., KB):", placeholder="XX", key="fab_initials").upper()
        fab_num_samples = st.number_input("Number of Samples to Generate:", min_value=1, value=1, step=1, key="fab_num_samples")

    st.markdown("---")
    st.subheader("Process Parameters & Details")
    
    # Use descriptive master name for the Excel sheet if available, otherwise use the ID.
    # The original notebook used reverse_mapping.get(int(master), "Unknown").
    # Here, fab_master_id is a string like "0", "1". We need to convert to int for lookup.
    try:
        master_id_int = int(fab_master_id)
        fab_master_name_for_excel = FAB_MASTER_NAME_DESCRIPTIVE_MAPPING.get(master_id_int, f"ID_{fab_master_id}")
    except ValueError:
        fab_master_name_for_excel = f"ID_{fab_master_id}" # Fallback if somehow not an int string
    
    # Display what will be written as "Master Name" in Excel for clarity
    st.caption(f"The 'Master Name' recorded in Excel for ID '{fab_master_id}' will be: '{fab_master_name_for_excel}'")


    param_col1, param_col2 = st.columns(2)
    with param_col1:
        fab_temperature = st.number_input("Temperature (°C):", value=FABRICATOR_DEFAULT_VALUES["Temperature"], format="%.1f", key="fab_temp")
        fab_pressure = st.number_input("Pressure (N):", value=FABRICATOR_DEFAULT_VALUES["Pressure"], format="%.1f", key="fab_pressure")
        fab_uv = st.number_input("UV (%):", value=FABRICATOR_DEFAULT_VALUES["UV"], format="%.1f", key="fab_uv")
        fab_uv_time = st.number_input("UV Time (s):", value=FABRICATOR_DEFAULT_VALUES["UV_Time"], format="%.1f", key="fab_uv_time") # Added UV Time
        fab_speed = st.number_input("Speed (m/s - in notebook, mm/s might be more typical for fab?):", value=FABRICATOR_DEFAULT_VALUES["Speed"], format="%.1f", key="fab_speed")
        fab_im_gap = st.number_input("Imprint Gap (mm):", value=FABRICATOR_DEFAULT_VALUES["Im_gap"], format="%.2f", key="fab_im_gap")
        fab_im_pressure = st.number_input("Imprint Pressure (N):", value=FABRICATOR_DEFAULT_VALUES["Im_pressure"], format="%.1f", key="fab_im_pressure")
        fab_del_gap = st.number_input("Delamination Gap (mm):", value=FABRICATOR_DEFAULT_VALUES["Del_gap"], format="%.2f", key="fab_del_gap")
        fab_del_pressure = st.number_input("Delamination Pressure (N):", value=FABRICATOR_DEFAULT_VALUES["Del_pressure"], format="%.1f", key="fab_del_pressure")

    with param_col2:
        fab_vacuum = st.number_input("Vacuum (mTor):", value=FABRICATOR_DEFAULT_VALUES["Vacuum"], format="%.1f", key="fab_vacuum")
        fab_pillar_pattern = st.selectbox("Pillar Pattern:", FAB_PILLAR_PATTERN, index=FAB_PILLAR_PATTERN.index(FABRICATOR_DEFAULT_VALUES["Pillar Pattern"]) if FABRICATOR_DEFAULT_VALUES["Pillar Pattern"] in FAB_PILLAR_PATTERN else 0, key="fab_pillar_pattern")
        fab_pillar_array = st.selectbox("Pillar Array:", FAB_PILLAR_ARRAY, index=FAB_PILLAR_ARRAY.index(FABRICATOR_DEFAULT_VALUES["Pillar Array"]) if FABRICATOR_DEFAULT_VALUES["Pillar Array"] in FAB_PILLAR_ARRAY else 0, key="fab_pillar_array")
        fab_primer = st.selectbox("Primer:", FAB_PRIMER, index=FAB_PRIMER.index(FABRICATOR_DEFAULT_VALUES["Primer"]) if FABRICATOR_DEFAULT_VALUES["Primer"] in FAB_PRIMER else 0, key="fab_primer")
        fab_pet = st.selectbox("PET:", FAB_PET, index=FAB_PET.index(FABRICATOR_DEFAULT_VALUES["PET"]) if FABRICATOR_DEFAULT_VALUES["PET"] in FAB_PET else 0, key="fab_pet")
        fab_metallisation = st.selectbox("Metallisation (True/False):", ["False", "True"], index=0 if FABRICATOR_DEFAULT_VALUES["Metallisation"] == "False" else 1, key="fab_metallisation")
        fab_metalised_material = st.text_input("Metalised Material (if Metallisation is True):", value=FABRICATOR_DEFAULT_VALUES["Metalised Material"], key="fab_metalised_material")
        fab_singulation = st.selectbox("Singulation (True/False):", ["False", "True"], index=0 if FABRICATOR_DEFAULT_VALUES["Singulation"] == "False" else 1, key="fab_singulation")
        fab_usability = st.selectbox("Usability (True/False):", ["False", "True"], index=0 if FABRICATOR_DEFAULT_VALUES["Usability"] == "False" else 1, key="fab_usability")

    fab_comments = st.text_area("Comments:", value=FABRICATOR_DEFAULT_VALUES["Comments"], height=100, key="fab_comments")
    
    # Default values not explicitly in UI for this tab but needed for append function
    fab_internal_name = FABRICATOR_DEFAULT_VALUES["Internal Name"] # Default from notebook
    fab_ips_name = "" # IPS Name - not in notebook UI, default to empty
    fab_anti_sticking2 = "" # Anti Sticking 2 - not in notebook UI, default to empty
    fab_no_of_prints = 0 # No of Prints - not in notebook UI, default to 0

    if st.button("Add Samples to Current Batch", key="fab_add_to_batch_button"):
        download_filename = st.session_state.fab_target_save_path
        if not download_filename:
            if st.session_state.fab_uploaded_excel_file:
                download_filename = st.session_state.fab_uploaded_excel_file.name
            else:
                download_filename = "generated_samples_batch.xlsx" 
        
        if not download_filename.lower().endswith(".xlsx"):
            download_filename += ".xlsx"

        selected_sheet_for_generation = st.session_state.fab_selected_sheet_name
        if not selected_sheet_for_generation:
            if not st.session_state.fab_manual_sheet_name_input:
                 selected_sheet_for_generation = "Sheet1"
            else:
                 selected_sheet_for_generation = st.session_state.fab_manual_sheet_name_input
        
        if not selected_sheet_for_generation:
            st.error("Please specify a Target Sheet Name.")
        elif not fab_initials or len(fab_initials) == 0:
            st.error("Please enter initials.")
        else:
            generated_names = generate_sample_name_fab(
                fab_material, fab_master_id, fab_salinisation, fab_anti_sticking,
                fab_resin, fab_resist, fab_initials, fab_num_samples
            )

            if generated_names:
                st.write("Adding to batch:")
                for name in generated_names:
                    st.text(name)
                
                # The append function now uses/updates the staged workbook or uploaded file
                # and returns the workbook object.
                modified_workbook = append_sample_data_to_excel_fab(
                    uploaded_file_obj=st.session_state.fab_uploaded_excel_file if not st.session_state.fab_staged_workbook_buffer else None, # Pass uploaded file only if no staged buffer
                    target_sheet_name=selected_sheet_for_generation,
                    sample_names_fab=generated_names,
                    internal_name_fab=fab_internal_name, 
                    material_fab=fab_material,
                    master_name_for_excel_fab=fab_master_name_for_excel,
                    ips_name_fab=fab_ips_name, 
                    anti_sticking_fab=fab_anti_sticking,
                    resin_fab=fab_resin,
                    anti_sticking2_fab=fab_anti_sticking2, 
                    resist_fab=fab_resist,
                    no_of_prints_fab=fab_no_of_prints, 
                    temperature_fab=fab_temperature,
                    pressure_fab=fab_pressure,
                    uv_fab=fab_uv,
                    uv_time_fab=fab_uv_time, 
                    speed_fab=fab_speed,
                    im_gap_fab=fab_im_gap,
                    im_pressure_fab=fab_im_pressure,
                    del_gap_fab=fab_del_gap,
                    del_pressure_fab=fab_del_pressure,
                    vacuum_fab=fab_vacuum,
                    pillar_pattern_fab=fab_pillar_pattern,
                    pillar_array_fab=fab_pillar_array,
                    primer_fab=fab_primer,
                    pet_fab=fab_pet,
                    metallisation_fab=fab_metallisation,
                    metalised_material_fab=fab_metalised_material,
                    singulation_fab=fab_singulation,
                    comments_fab=fab_comments,
                    usability_fab=fab_usability
                )
                
                if modified_workbook:
                    try:
                        # Save workbook to an in-memory buffer for staging
                        excel_buffer = io.BytesIO()
                        modified_workbook.save(excel_buffer)
                        # excel_buffer.seek(0) # No need to seek here, will be done before reading
                        st.session_state.fab_staged_workbook_buffer = excel_buffer
                        
                        # Add generated names to the staged list (avoiding duplicates if re-generating same)
                        for name in generated_names:
                            if name not in st.session_state.fab_staged_sample_names:
                                st.session_state.fab_staged_sample_names.append(name)
                        
                        st.success(f"{len(generated_names)} sample(s) added to the current batch!")
                        # Do not show download button here, it will be separate
                    except Exception as e:
                        st.error(f"Error adding samples to batch: {e}")
            else:
                st.info("No sample names were generated. Check input parameters.")

    st.markdown("---")
    st.subheader("Current Batch")

    if st.session_state.fab_staged_sample_names:
        st.write("Samples currently in batch:")
        for name in st.session_state.fab_staged_sample_names:
            st.text(f"- {name}")
        
        # Determine the download filename for the batch
        batch_download_filename = st.session_state.fab_target_save_path
        if not batch_download_filename:
            if st.session_state.fab_uploaded_excel_file and not st.session_state.fab_staged_workbook_buffer: # If using uploaded as template for first batch
                batch_download_filename = st.session_state.fab_uploaded_excel_file.name
            else: # Default name if no path given or if buffer already exists
                batch_download_filename = "fabricated_samples_batch.xlsx"
        if not batch_download_filename.lower().endswith(".xlsx"):
            batch_download_filename += ".xlsx"

        if st.session_state.fab_staged_workbook_buffer and st.session_state.fab_staged_workbook_buffer.getvalue():
            st.download_button(
                label="📥 Download Batch File",
                data=st.session_state.fab_staged_workbook_buffer.getvalue(), # Getvalue directly here
                file_name=batch_download_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="fab_download_batch_button"
            )
        else:
            st.warning("No batch data to download. Add samples first.")

        if st.button("Clear Current Batch", key="fab_clear_batch_button"):
            st.session_state.fab_staged_workbook_buffer = None
            st.session_state.fab_staged_sample_names = []
            st.session_state.fab_uploaded_excel_file = None # Also clear uploaded file if batch is cleared, to start fresh
            st.session_state.fab_target_save_path = "" # Clear suggested save path
            st.session_state.fab_selected_sheet_name = None 
            st.session_state.fab_excel_sheets_options = []
            st.session_state.fab_df_preview = None
            st.info("Current batch cleared.")
            st.rerun() # Rerun to update UI
    else:
        st.info("No samples added to the batch yet.")

# New Rclone Tab
with tab_rclone:
    st.subheader("SharePoint File Downloader (Rclone)")
    st.markdown("""
    **Important:** 
    1.  `rclone` must be installed on the system running this Streamlit application.
    2.  The `rclone` executable must be accessible via the system's PATH environment variable, OR you can provide the full path to `rclone.exe` below.
    3.  A SharePoint remote must be configured in your `rclone.conf` file (e.g., named `MySharePoint:`).
    """)

    st.session_state.rclone_exe_path = st.text_input(
        "Path to rclone.exe (optional, defaults to 'rclone' if in PATH):",
        value=st.session_state.rclone_exe_path,
        key="rclone_exe_path_input",
        placeholder="e.g., C:\\path\\to\\rclone.exe or leave blank if rclone is in PATH"
    )
    # Ensure session state is updated from text input
    # st.session_state.rclone_exe_path = st.session_state.rclone_exe_path_input # This line is removed
    # Default to "rclone" if input is cleared by user, to ensure it's never truly empty for command construction
    # This defaulting will now happen inside the button logic if needed.

    st.session_state.rclone_remote_name = st.text_input(
        "Rclone SharePoint Remote Name:", 
        value=st.session_state.rclone_remote_name,
        key="rclone_remote_name_input",
        placeholder="e.g., MySharePoint"
    )

    st.session_state.rclone_source_path = st.text_input(
        "SharePoint Source Path (File or Folder):", 
        value=st.session_state.rclone_source_path,
        key="rclone_source_path_input",
        placeholder="e.g., Shared Documents/ProjectX/report.docx or Shared Documents/ProjectX/"
    )

    st.session_state.rclone_local_destination = st.text_input(
        "Local Destination Directory:", 
        value=st.session_state.rclone_local_destination,
        key="rclone_local_dest_input",
        placeholder="e.g., C:\\Users\\YourName\\Downloads\\ProjectX"
    )
    # Update session state immediately from text input if changed
    st.session_state.rclone_remote_name = st.session_state.rclone_remote_name_input
    st.session_state.rclone_source_path = st.session_state.rclone_source_path_input
    st.session_state.rclone_local_destination = st.session_state.rclone_local_dest_input

    # --- Local Destination Directory Browser ---
    def rclone_update_dest_from_text_input():
        st.session_state.rclone_local_destination = st.session_state.rclone_local_dest_input_val
        # If browser is open and path changes, try to sync browser path if it's a valid dir
        if st.session_state.rclone_show_local_dest_browser and os.path.isdir(st.session_state.rclone_local_dest_input_val):
            st.session_state.rclone_local_dest_browser_path = st.session_state.rclone_local_dest_input_val
    
    # Use a different key for the text input that the browser controls
    st.session_state.rclone_local_dest_input_val = st.text_input(
        "Local Destination Directory:", 
        value=st.session_state.rclone_local_destination, # This now refers to the main session state
        key="rclone_local_dest_input_val_key",
        on_change=rclone_update_dest_from_text_input,
        placeholder="e.g., C:\\Users\\YourName\\Downloads\\ProjectX"
    )
    st.session_state.rclone_local_destination = st.session_state.rclone_local_dest_input_val # Ensure main state is updated

    if st.button("Browse / Hide Local Destination Browser", key="rclone_toggle_local_dest_browser_btn"):
        st.session_state.rclone_show_local_dest_browser = not st.session_state.rclone_show_local_dest_browser
        if st.session_state.rclone_show_local_dest_browser:
            # Initialize browser path to current destination if valid, else home
            if os.path.isdir(st.session_state.rclone_local_destination):
                st.session_state.rclone_local_dest_browser_path = st.session_state.rclone_local_destination
            else:
                st.session_state.rclone_local_dest_browser_path = os.path.expanduser("~")

    if st.session_state.rclone_show_local_dest_browser:
        with st.expander("Local Destination Browser", expanded=True):
            st.write(f"Browsing: `{st.session_state.rclone_local_dest_browser_path}`")

            if st.button("⬆️ Up One Level", key="rclone_browse_dest_up_btn"):
                parent = str(Path(st.session_state.rclone_local_dest_browser_path).parent)
                if os.path.isdir(parent) and parent != st.session_state.rclone_local_dest_browser_path:
                    st.session_state.rclone_local_dest_browser_path = parent
                    st.rerun()
            
            browser_dirs, _ = list_directory_contents(st.session_state.rclone_local_dest_browser_path)
            if browser_dirs:
                st.write("Subdirectories (click to navigate):")
                for i, dir_item in enumerate(browser_dirs):
                    dir_path_str = str(dir_item.resolve())
                    clean_dir_name = "".join(c if c.isalnum() or c in ['_'] else '_' for c in dir_item.name)
                    button_key = f"rclone_browse_dest_nav_to_dir_idx_{i}_{clean_dir_name}"
                    if st.button(f"📁 {dir_item.name}", key=button_key):
                        st.session_state.rclone_local_dest_browser_path = dir_path_str
                        st.rerun()
            else:
                st.write("No subdirectories in current browsing path.")

            select_current_label = f"Use '{os.path.basename(st.session_state.rclone_local_dest_browser_path)}' as Destination"
            if st.button(select_current_label, key="rclone_browse_dest_select_current_btn"):
                st.session_state.rclone_local_destination = st.session_state.rclone_local_dest_browser_path
                st.session_state.rclone_local_dest_input_val = st.session_state.rclone_local_destination # Update text field
                st.session_state.rclone_show_local_dest_browser = False 
                st.rerun() 
            
            if st.button("Cancel Browsing", key="rclone_cancel_browsing_dest_btn"):
                st.session_state.rclone_show_local_dest_browser = False
                st.rerun()
    # --- End Local Destination Directory Browser ---

    if st.button("Download with Rclone", key="rclone_download_button"):
        if not st.session_state.rclone_remote_name:
            st.error("Please enter the Rclone SharePoint Remote Name.")
        elif not st.session_state.rclone_source_path:
            st.error("Please enter the SharePoint Source Path.")
        elif not st.session_state.rclone_local_destination:
            st.error("Please enter the Local Destination Directory.")
        elif not os.path.isdir(st.session_state.rclone_local_destination):
            st.error(f"Local Destination Directory is not valid or does not exist: {st.session_state.rclone_local_destination}")
        else:
            # Explicitly update rclone_exe_path from the input field before using it
            current_exe_path_input = st.session_state.rclone_exe_path_input.strip()
            if not current_exe_path_input:
                effective_rclone_exe_path = "rclone" # Default if empty
            else:
                effective_rclone_exe_path = current_exe_path_input
            st.session_state.rclone_exe_path = effective_rclone_exe_path # Store the used path

            # Quote paths for command string if they contain spaces
            quoted_rclone_exe = f'"{effective_rclone_exe_path}"' if ' ' in effective_rclone_exe_path and not effective_rclone_exe_path.startswith('"') else effective_rclone_exe_path
            
            # Remote path: remote_name:"path part"
            # The path part of the remote needs to be quoted if it contains spaces.
            # rclone itself handles the remote:path syntax.
            # We are creating a string that rclone CLI will parse.
            rclone_source_remote = st.session_state.rclone_remote_name.strip().rstrip(":")
            rclone_source_path_part = st.session_state.rclone_source_path.strip()
            # No, rclone needs quotes around the whole "remote:path" if the path part has spaces for cmd.exe
            # Let's try quoting the path part for rclone's parsing first: remote:"path"
            # then if that fails, quote the whole thing for cmd.exe
            
            # For the batch file, it's safer to quote the arguments that rclone receives.
            # rclone.exe copy remote:"path with space" "local path with space"
            
            quoted_source_path_part = f'"{rclone_source_path_part}"' # Always quote for simplicity, rclone handles it.
            full_remote_source_for_rclone = f'{rclone_source_remote}:{quoted_source_path_part}'
            
            local_dest_cleaned = st.session_state.rclone_local_destination.strip()
            quoted_local_dest = f'"{local_dest_cleaned}"' if ' ' in local_dest_cleaned and not local_dest_cleaned.startswith('"') else local_dest_cleaned

            # Construct the rclone command that will go INSIDE the batch file
            rclone_command_for_batch_file = (
                f'{quoted_rclone_exe} copy -v --create-empty-src-dirs '
                f'{full_remote_source_for_rclone} {quoted_local_dest}'
            )
            
            st.session_state.rclone_command_output = f"Preparing batch file...\n{rclone_command_for_batch_file}" # Show what will be in batch

            try:
                # Create a temporary batch file
                with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.bat', encoding='utf-8') as tmp_batch_file:
                    tmp_batch_file.write(rclone_command_for_batch_file)
                    st.session_state.rclone_temp_batch_file_path = tmp_batch_file.name
                
                # Command to execute the batch file
                # Quote the batch file path in case the temp directory has spaces
                command_to_run_batch = f'cmd /c "{st.session_state.rclone_temp_batch_file_path}"'
                st.session_state.rclone_last_run_command = command_to_run_batch
                
                st.session_state.rclone_command_output = f"""Attempting to execute rclone command via temporary batch file: {st.session_state.rclone_temp_batch_file_path}
Batch file content:
{rclone_command_for_batch_file}
Executing: {command_to_run_batch}"""
                st.session_state.rclone_is_running = True
                st.rerun()

            except Exception as e:
                st.error(f"Error creating temporary batch file: {str(e)}")
                st.session_state.rclone_is_running = False


    # Display spinner if rclone is supposed to be running 
    if st.session_state.get('rclone_is_running', False):
        with st.spinner("Rclone operation in progress via batch file..."):
            # The actual run_terminal_cmd will be called by the agent after this rerun
            # based on rclone_is_running and rclone_last_run_command being set.
            pass 

    st.markdown("---")
    st.subheader("Rclone Output")
    if st.session_state.rclone_last_run_command:
        st.caption(f"Last command attempted: `{st.session_state.rclone_last_run_command}`")
    st.text_area("Output:", value=st.session_state.rclone_command_output, height=200, disabled=True, key="rclone_output_area")


# Tab 3: Excel Row Exporter
with tab3:
    st.subheader("Excel Row Exporter")
    
    uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xls'])
    
    if uploaded_file is not None:
        try:
            # Read Excel file
            excel_file = pd.ExcelFile(uploaded_file)
            sheet_name = st.selectbox("Select Sheet", excel_file.sheet_names)
            
            # Read the selected sheet
            df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
            
            # Show DataFrame preview
            st.write("Preview of the Excel file:")
            st.dataframe(df.head())
            
            # Input for row range
            col1, col2 = st.columns(2)
            with col1:
                start_row = st.number_input("Start Row", min_value=1, max_value=len(df), value=1)
            with col2:
                end_row = st.number_input("End Row", min_value=1, max_value=len(df), value=min(5, len(df)))
            
            # Output folder
            output_folder = st.text_input("Output Folder Path", placeholder="Enter the folder path for CSV files", key="excel_folder")
            
            if st.button("Export Rows"):
                if not output_folder:
                    st.warning("Please enter an output folder path.")
                elif not os.path.exists(output_folder):
                    st.error("Output folder does not exist.")
                else:
                    saved_files = save_rows_as_csv(df, output_folder, start_row, end_row)
                    if saved_files:
                        st.success(f"Successfully exported {len(saved_files)} files!")
                        st.write("Exported files:")
                        for file_name in saved_files:
                            st.text(f"- {file_name}.csv")
                    else:
                        st.warning("No files were exported. Please check your row range and data.")
                        
        except Exception as e:
            st.error(f"Error reading Excel file: {str(e)}")

# Tab 4: Data Structure Creator
with tab4:
    st.subheader("Data Structure Creator")
    
    # Two methods: Excel file or manual Sample ID
    method = st.radio("Choose Method", ["Upload Excel File", "Enter Sample ID Manually"])
    
    col1, col2 = st.columns(2)
    with col1:
        fabrication = st.checkbox("Include Fabrication Folder")
    with col2:
        inspection = st.checkbox("Include Inspection Folders")
    
    if method == "Upload Excel File":
        excel_file = st.file_uploader("Choose Excel File", type=['xlsx', 'xls'], key="structure_excel")
        
        if excel_file is not None:
            # Read Excel file
            excel_data = pd.ExcelFile(excel_file)
            sheet_name = st.selectbox("Select Sheet", excel_data.sheet_names, key="structure_sheet")
            
            # Read the selected sheet
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Show DataFrame preview
            st.write("Preview of the Excel file:")
            st.dataframe(df.head())
            
            # Row range selection
            col1, col2 = st.columns(2)
            with col1:
                start_row = st.number_input("Start Row", min_value=1, max_value=len(df), value=1, key="structure_start")
            with col2:
                end_row = st.number_input("End Row", min_value=1, max_value=len(df), value=min(5, len(df)), key="structure_end")
    else:
        sample_id = st.text_input("Enter Sample ID")
    
    # --- Output Location with Browse ---
    if 'save_location_val' not in st.session_state:
        st.session_state.save_location_val = os.path.expanduser("~")
    if 'output_location_browser_path' not in st.session_state:
        st.session_state.output_location_browser_path = st.session_state.save_location_val 
    if 'show_dir_browser_tab4' not in st.session_state: 
        st.session_state.show_dir_browser_tab4 = False

    # Callback to update the main session state variable from text input
    def update_save_location_from_text_input_tab4():
        if "structure_save_location_text_input_key_tab4" in st.session_state:
            st.session_state.save_location_val = st.session_state.structure_save_location_text_input_key_tab4

    st.text_input(
        "Output Location",
        value=st.session_state.save_location_val,
        key="structure_save_location_text_input_key_tab4",
        on_change=update_save_location_from_text_input_tab4,
        placeholder="Enter path or browse below"
    )

    if st.button("Browse / Hide Browser", key="browse_toggle_tab4_btn"):
        st.session_state.show_dir_browser_tab4 = not st.session_state.show_dir_browser_tab4
        if st.session_state.show_dir_browser_tab4:
            # Initialize browser path to current save_location_val if it's a dir, else home
            current_potential_path = st.session_state.save_location_val
            if os.path.isdir(current_potential_path):
                st.session_state.output_location_browser_path = current_potential_path
            else:
                st.session_state.output_location_browser_path = os.path.expanduser("~")

    if st.session_state.show_dir_browser_tab4:
        with st.expander("Directory Browser", expanded=True):
            st.write(f"Current browsing path: `{st.session_state.output_location_browser_path}`")

            if st.button("⬆️ Up One Level", key="output_loc_browse_up_tab4_btn"):
                parent = str(Path(st.session_state.output_location_browser_path).parent)
                if os.path.isdir(parent) and parent != st.session_state.output_location_browser_path:
                    st.session_state.output_location_browser_path = parent
                    st.rerun()

            browser_dirs, _ = list_directory_contents(st.session_state.output_location_browser_path)
            if browser_dirs:
                st.write("Subdirectories (click to navigate):")
                for i, dir_item in enumerate(browser_dirs):
                    dir_path_str = str(dir_item.resolve()) 
                    # Sanitize dir_item.name for the key and add index for uniqueness
                    clean_dir_name = "".join(c if c.isalnum() or c in ['_'] else '_' for c in dir_item.name)
                    button_key = f"browse_nav_to_dir_tab4_idx_{i}_{clean_dir_name}"
                    if st.button(f"📁 {dir_item.name}", key=button_key):
                        st.session_state.output_location_browser_path = dir_path_str
                        st.rerun()
            else:
                st.write("No subdirectories in current browsing path.")

            select_current_label = f"Use '{os.path.basename(st.session_state.output_location_browser_path)}' as Output Location"
            if st.button(select_current_label, key="output_loc_browse_select_current_tab4_btn"):
                st.session_state.save_location_val = st.session_state.output_location_browser_path
                st.session_state.show_dir_browser_tab4 = False 
                st.rerun() 
            
            if st.button("Cancel Browsing", key="cancel_browsing_tab4_btn"):
                st.session_state.show_dir_browser_tab4 = False
                st.rerun()
    # --- End Output Location with Browse ---
    
    if st.button("Create Folder Structure"):
        final_save_location = st.session_state.save_location_val 
        if not final_save_location:
            st.error("Please specify an output location")
        elif not os.path.isdir(final_save_location): 
            st.error(f"Output location is not a valid directory or does not exist: {final_save_location}")
        else:
            if method == "Upload Excel File" and excel_file is not None:
                success_count = 0
                error_messages = []
                
                # Process Excel rows
                for index in range(start_row-1, end_row):
                    row = df.iloc[index]
                    if not row.isnull().all():
                        sample_id = str(row[0])
                        if pd.notna(sample_id) and sample_id:
                            success, result = create_folders_for_csv(
                                f"{sample_id}.csv",
                                sample_id,
                                pd.DataFrame([row.values], columns=df.columns),
                                final_save_location,
                                fabrication,
                                inspection
                            )
                            if success:
                                success_count += 1
                            else:
                                error_messages.append(f"Error with {sample_id}: {result}")
                
                if success_count > 0:
                    st.success(f"Successfully created folder structures for {success_count} samples!")
                if error_messages:
                    for msg in error_messages:
                        st.error(msg)
                        
            else:  # Manual Sample ID
                if not sample_id:
                    st.error("Please enter a Sample ID")
                else:
                    success, result = create_folders_for_csv(
                        f"{sample_id}.csv",
                        sample_id,
                        pd.DataFrame(),
                        final_save_location,
                        fabrication,
                        inspection
                    )
                    if success:
                        st.success(f"Successfully created folder structure for {sample_id}!")
                        st.write("Created folders:")
                        for folder in result:
                            st.text(f"📁 {os.path.basename(folder)}")
                    else:
                        st.error(f"Error creating folders: {result}")

# Tab 5: Line Scaling
with tab5:
    st.subheader("Line Scaling Tool")
    st.write("Scale and visualize lines based on different working areas")

    # Input for working areas
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Original Working Area")
        old_width = st.number_input("Original Width (mm)", value=1300, step=50, key="old_width_ls")
        old_height = st.number_input("Original Height (mm)", value=1100, step=50, key="old_height_ls")
    
    with col2:
        st.subheader("New Working Area")
        new_width = st.number_input("New Width (mm)", value=650, step=50, key="new_width_ls")
        new_height = st.number_input("New Height (mm)", value=550, step=50, key="new_height_ls")

    # Create input for lines and parameters
    st.subheader("Line Parameters")
    
    # Prepare data for the editor
    line_editor_data = []
    for i, line_coords in enumerate(st.session_state.lines):
        line_color = PLOT_COLORS[i % len(PLOT_COLORS)]
        line_editor_data.append({
            "Line": f"Line {i+1}",
            "X start (mm)": float(line_coords[0][0]),
            "Y start (mm)": float(line_coords[0][1]),
            "X end (mm)": float(line_coords[1][0]),
            "Y end (mm)": float(line_coords[1][1]),
            "Speed (mm/s)": float(st.session_state.speed[i]),
            "T cycle (ms)": int(st.session_state.t_cycle[i]),
            "T pulse (ms)": int(st.session_state.t_pulse[i]),
            "Color": line_color
        })
    
    df_lines_editable = pd.DataFrame(line_editor_data)
    
    # Use st.data_editor to allow editing
    edited_df = st.data_editor(
        df_lines_editable, 
        num_rows="dynamic", # Allow adding/deleting rows
        key="line_param_editor",
        column_config={
            "Line": st.column_config.TextColumn(disabled=True), # Make Line column non-editable
             "X start (mm)": st.column_config.NumberColumn(format="%.2f", required=True),
             "Y start (mm)": st.column_config.NumberColumn(format="%.2f", required=True),
             "X end (mm)": st.column_config.NumberColumn(format="%.2f", required=True),
             "Y end (mm)": st.column_config.NumberColumn(format="%.2f", required=True),
             "Speed (mm/s)": st.column_config.NumberColumn(format="%.2f", required=True),
             "T cycle (ms)": st.column_config.NumberColumn(format="%d", required=True),
             "T pulse (ms)": st.column_config.NumberColumn(format="%d", required=True),
             "Color": st.column_config.TextColumn(disabled=True) # Make Color column non-editable
        }
    )

    # Update session state from the edited DataFrame
    if edited_df is not None and not edited_df.equals(df_lines_editable): # Check if changes were made
        new_lines = []
        new_speeds = []
        new_t_cycles = []
        new_t_pulses = []
        
        for index, row in edited_df.iterrows():
            try:
                new_lines.append(
                    ((float(row["X start (mm)"]), float(row["Y start (mm)"])), 
                     (float(row["X end (mm)"]), float(row["Y end (mm)"])))
                )
                new_speeds.append(float(row["Speed (mm/s)"]))
                new_t_cycles.append(int(row["T cycle (ms)"]))
                new_t_pulses.append(int(row["T pulse (ms)"]))
            except KeyError as e:
                st.error(f"Missing column in edited data: {e}. Please ensure all columns are present.")
                # Prevent further processing if a column is missing (e.g., after row deletion)
                st.stop()
            except ValueError as e:
                st.error(f"Invalid data type for a parameter: {e}. Please enter valid numbers.")
                st.stop()


        st.session_state.lines = new_lines
        st.session_state.speed = new_speeds
        st.session_state.t_cycle = new_t_cycles
        st.session_state.t_pulse = new_t_pulses
        # st.rerun() # Rerun to reflect changes immediately if needed, or rely on button press

    if st.button("Scale Lines and Visualize", key="scale_visualize_button"):
        if not st.session_state.lines:
            st.warning("No lines to scale. Please add lines in the table above.")
        else:
            # Scale the lines
            old_area = (old_width, old_height)
            new_area = (new_width, new_height)
            
            try:
                scaled_lines = scale_lines(st.session_state.lines, old_area, new_area)
            except Exception as e:
                st.error(f"Error during scaling: {e}")
                st.stop()

            # Create visualization
            col_viz1, col_viz2 = st.columns(2)
            
            with col_viz1:
                st.subheader("Original Lines")
                fig_original = draw_lines(st.session_state.lines, old_area, title="Original Lines", colors_list=PLOT_COLORS)
                st.pyplot(fig_original)
                
            with col_viz2:
                st.subheader("Scaled Lines")
                fig_scaled = draw_lines(scaled_lines, new_area, title="Scaled Lines", colors_list=PLOT_COLORS)
                st.pyplot(fig_scaled)

            # Display scaled results
            st.subheader("Scaled Line Parameters")
            
            scaled_results_data = []
            for i, (original_line, scaled_line_coords) in enumerate(zip(st.session_state.lines, scaled_lines)):
                line_color = PLOT_COLORS[i % len(PLOT_COLORS)]
                scaled_results_data.append({
                    "Line": f"Line {i+1}",
                    "Original X start": round(original_line[0][0], 2),
                    "Original Y start": round(original_line[0][1], 2),
                    "Original X end": round(original_line[1][0], 2),
                    "Original Y end": round(original_line[1][1], 2),
                    "Scaled X start": round(scaled_line_coords[0][0], 2),
                    "Scaled Y start": round(scaled_line_coords[0][1], 2),
                    "Scaled X end": round(scaled_line_coords[1][0], 2),
                    "Scaled Y end": round(scaled_line_coords[1][1], 2),
                    "Speed (mm/s)": round(st.session_state.speed[i], 2),
                    "T cycle (ms)": st.session_state.t_cycle[i],
                    "T pulse (ms)": st.session_state.t_pulse[i],
                    "Color": line_color
                })
            df_scaled_results = pd.DataFrame(scaled_results_data)
            
            # Apply styling to highlight out-of-bounds coordinates and format numbers
            if not df_scaled_results.empty:
                styled_df = df_scaled_results.style.apply(
                    highlight_out_of_bounds_styler, 
                    new_w=new_width, 
                    new_h=new_height, 
                    axis=1,
                    subset=['Scaled X start', 'Scaled Y start', 'Scaled X end', 'Scaled Y end']
                ).format({
                    'Original X start': '{:.2f}',
                    'Original Y start': '{:.2f}',
                    'Original X end': '{:.2f}',
                    'Original Y end': '{:.2f}',
                    'Scaled X start': '{:.2f}',
                    'Scaled Y start': '{:.2f}',
                    'Scaled X end': '{:.2f}',
                    'Scaled Y end': '{:.2f}',
                    'Speed (mm/s)': '{:.2f}'
                })
                st.dataframe(styled_df)
            else:
                st.dataframe(df_scaled_results) # Show empty dataframe if no results

# Add a footer with timestamp
st.markdown("---")
st.markdown(f"Last updated: {time.strftime('%Y-%m-%d %H:%M:%S')}")

# Auto-refresh functionality
if st.button("Refresh Data"):
    st.experimental_rerun() 